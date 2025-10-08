import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import pandas as pd
from PIL import Image, ImageTk
import webbrowser
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, NamedStyle

class ProExcelMergerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Pro Merger v2.0")

        # File paths
        self.sales_journal_path = tk.StringVar()
        self.job_master_path = tk.StringVar()
        self.job_ledger_path = tk.StringVar()
        self.job_estimates_path = tk.StringVar()
        self.output_file_path = tk.StringVar()

        self.key_columns = ['Job ID', 'Cost Code ID', 'Phase ID']
        self.required_column = None
        self.original_columns = []
        self.rename_map = {}
        self.calculation_columns = {}
        self.phase_id_criteria = 'Supervision-Wages' # Default value
        self.config_file = 'merger_config_v2.json'
        self.version = "2.0.0"

        self.load_app_configuration()
        self.create_menu()

        # Main frame
        self.main_frame = tk.Frame(root, padx=10, pady=10)
        self.main_frame.pack(padx=10, pady=10)

        # Add logo
        self.set_logo("C:/Users/dale/OneDrive/Documents/ISI/BnL/Sage/Images/BnL.Logo.jpg", self.main_frame)

        # File selection UI
        row_num = 1
        # Sales Journal
        tk.Label(self.main_frame, text="Custom Sales Journal:").grid(row=row_num, column=0, padx=5, pady=5, sticky='w')
        tk.Entry(self.main_frame, textvariable=self.sales_journal_path, width=50, state='readonly').grid(row=row_num, column=1, padx=5, pady=5)
        tk.Button(self.main_frame, text="Browse...", command=self.browse_sales_journal).grid(row=row_num, column=2, padx=5, pady=5)
        row_num += 1

        # Job Master File
        tk.Label(self.main_frame, text="Custom Job Master File:").grid(row=row_num, column=0, padx=5, pady=5, sticky='w')
        tk.Entry(self.main_frame, textvariable=self.job_master_path, width=50, state='readonly').grid(row=row_num, column=1, padx=5, pady=5)
        tk.Button(self.main_frame, text="Browse...", command=self.browse_job_master).grid(row=row_num, column=2, padx=5, pady=5)
        row_num += 1

        # Job Ledger
        tk.Label(self.main_frame, text="Custom Job Ledger - A:").grid(row=row_num, column=0, padx=5, pady=5, sticky='w')
        tk.Entry(self.main_frame, textvariable=self.job_ledger_path, width=50, state='readonly').grid(row=row_num, column=1, padx=5, pady=5)
        tk.Button(self.main_frame, text="Browse...", command=self.browse_job_ledger).grid(row=row_num, column=2, padx=5, pady=5)
        row_num += 1

        # Job Estimates
        tk.Label(self.main_frame, text="Estimated Job Expenses - A:").grid(row=row_num, column=0, padx=5, pady=5, sticky='w')
        tk.Entry(self.main_frame, textvariable=self.job_estimates_path, width=50, state='readonly').grid(row=row_num, column=1, padx=5, pady=5)
        tk.Button(self.main_frame, text="Browse...", command=self.browse_job_estimates).grid(row=row_num, column=2, padx=5, pady=5)
        row_num += 1

        # Output file path
        tk.Label(self.main_frame, text="Output File:").grid(row=row_num, column=0, padx=5, pady=5, sticky='w')
        tk.Entry(self.main_frame, textvariable=self.output_file_path, width=50, state='readonly').grid(row=row_num, column=1, padx=5, pady=5)
        tk.Button(self.main_frame, text="Browse...", command=self.browse_output_file).grid(row=row_num, column=2, padx=5, pady=5)

        # Go button
        self.go_button = tk.Button(root, text="Go", command=self.go_process, state=tk.DISABLED)
        self.go_button.pack(pady=10)

    def create_menu(self):
        self.menubar = tk.Menu(self.root)
        
        self.filemenu = tk.Menu(self.menubar, tearoff=0)
        self.filemenu.add_command(label="Select Logo", command=self.select_logo_config)
        self.filemenu.add_command(label="Configure Columns", command=self.open_column_config)
        self.filemenu.add_separator()
        self.filemenu.add_command(label="Exit", command=self.root.quit)
        self.menubar.add_cascade(label="File", menu=self.filemenu)

        helpmenu = tk.Menu(self.menubar, tearoff=0)
        helpmenu.add_command(label="About", command=self.show_about)
        self.menubar.add_cascade(label="Help", menu=helpmenu)

        self.root.config(menu=self.menubar)

    def load_app_configuration(self):
        try:
            with open(self.config_file, 'r') as f:
                config = json.load(f)
                self.phase_id_criteria = config.get('phase_id_criteria', 'Supervision-Wages')
        except FileNotFoundError:
            pass # Keep default if file not found

    def save_app_configuration(self):
        config = {'phase_id_criteria': self.phase_id_criteria}
        with open(self.config_file, 'w') as f:
            json.dump(config, f, indent=4)

    def open_column_config(self):
        config_window = tk.Toplevel(self.root)
        config_window.title("Configure Columns")

        tk.Label(config_window, text="Select Phase ID for Percent Complete calculation:").pack(pady=10)

        phase_ids = ['Supervision-Wages'] # Default
        if self.job_estimates_path.get():
            try:
                df = pd.read_excel(self.job_estimates_path.get())
                if 'Phase ID' in df.columns:
                    phase_ids = df['Phase ID'].dropna().unique().tolist()
            except Exception as e:
                messagebox.showerror("Error", f"Could not load Phase IDs: {e}", parent=config_window)

        selected_phase_id = tk.StringVar(value=self.phase_id_criteria)
        option_menu = tk.OptionMenu(config_window, selected_phase_id, *phase_ids)
        option_menu.pack(pady=10, padx=10)

        def save_and_close():
            self.phase_id_criteria = selected_phase_id.get()
            self.save_app_configuration()
            messagebox.showinfo("Saved", "Configuration saved successfully.", parent=config_window)
            config_window.destroy()

        tk.Button(config_window, text="Save & Close", command=save_and_close).pack(pady=10)

    def show_about(self):
        about_window = tk.Toplevel(self.root)
        about_window.title("About Pro Excel Merger")
        
        logo_path = "C:/Users/dale/OneDrive/Documents/ISI/BnL/Sage/Images/IDS.Logo.png"
        logo_image = Image.open(logo_path)
        logo_image = logo_image.resize((100, 100), Image.LANCZOS)
        self.about_logo_photo = ImageTk.PhotoImage(logo_image)
        
        tk.Label(about_window, image=self.about_logo_photo).pack(pady=10)
        tk.Label(about_window, text=f"Version: {self.version}").pack(pady=5)
        
        tk.Label(about_window, text="Contact: Dale Linn").pack(pady=5)
        email = tk.Label(about_window, text="richard.linn@biz-solve.com", fg="blue", cursor="hand2")
        email.pack()
        email.bind("<Button-1>", lambda e: webbrowser.open_new("mailto:richard.linn@biz-solve.com"))

        tk.Label(about_window, text="For new versions, please visit:").pack(pady=5)
        link = tk.Label(about_window, text="https://github.com/dlerhetal/BnL.ExcelMerge/releases", fg="blue", cursor="hand2")
        link.pack()
        link.bind("<Button-1>", lambda e: webbrowser.open_new("https://github.com/dlerhetal/BnL.ExcelMerge/releases"))

    def set_logo(self, logo_path, frame):
        try:
            logo_image = Image.open(logo_path)
            aspect_ratio = logo_image.width / logo_image.height
            logo_image = logo_image.resize((int(50 * aspect_ratio), 50), Image.LANCZOS)
            self.logo_photo = ImageTk.PhotoImage(logo_image)
            
            if hasattr(self, 'logo_label'):
                self.logo_label.config(image=self.logo_photo)
            else:
                self.logo_label = tk.Label(frame, image=self.logo_photo)
                self.logo_label.grid(row=0, column=0, columnspan=3, pady=10)

        except Exception as e:
            messagebox.showerror("Error", f"Could not load logo: {e}")

    def select_logo_config(self):
        filename = filedialog.askopenfilename(filetypes=[
            ("Image files", "*.jpg *.jpeg *.png *.gif *.bmp"),
            ("All files", "*.*",)
        ])
        if filename:
            self.set_logo(filename, self.main_frame)

    def browse_sales_journal(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if filename:
            self.sales_journal_path.set(filename)
            self.check_inputs()

    def browse_job_master(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if filename:
            self.job_master_path.set(filename)
            self.check_inputs()

    def browse_job_ledger(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if filename:
            self.job_ledger_path.set(filename)
            self.check_inputs()

    def browse_job_estimates(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if filename:
            self.job_estimates_path.set(filename)
            self.check_inputs()

    def browse_output_file(self):
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if filename:
            self.output_file_path.set(filename)
            self.check_inputs()

    def check_inputs(self, *args):
        if (self.sales_journal_path.get() and
            self.job_master_path.get() and
            self.job_ledger_path.get() and
            self.job_estimates_path.get() and
            self.output_file_path.get()):
            self.go_button.config(state=tk.NORMAL)
            # self.filemenu.entryconfig("Configure Columns", state=tk.NORMAL) # Disabled for now
        else:
            self.go_button.config(state=tk.DISABLED)
            # if hasattr(self, 'filemenu'):
                # self.filemenu.entryconfig("Configure Columns", state=tk.DISABLED)

    def clean_job_ledger(self, df):
        # Forward fill missing values in specified columns
        df['Job ID'].fillna(method='ffill', inplace=True)
        df['Cost Code ID'].fillna(method='ffill', inplace=True)
        df['Phase Description'].fillna(method='ffill', inplace=True)
        df['Phase ID'].fillna(method='ffill', inplace=True)

        # Drop rows where 'Trans Description' is NaN, which are likely empty or summary rows
        df.dropna(subset=['Trans Description'], inplace=True)

        # Remove rows containing 'Total' in 'Job ID' or 'Trans Description'
        df = df[~df['Job ID'].astype(str).str.contains('Total', na=False)]
        df = df[~df['Job ID'].astype(str).str.contains('Report', na=False)]
        df = df[~df['Trans Description'].astype(str).str.contains('Total', na=False)]

        return df

    def to_numeric_safe(self, df, columns):
        for col in columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col].astype(str).str.replace('[$,]', '', regex=True), errors='coerce').fillna(0)
        return df

    def go_process(self):
        try:
            df_sales_journal = pd.read_excel(self.sales_journal_path.get())
            df_job_master = pd.read_excel(self.job_master_path.get())
            df_job_ledger = pd.read_excel(self.job_ledger_path.get())
            df_job_estimates = pd.read_excel(self.job_estimates_path.get())

            # Clean job estimates
            df_job_estimates.dropna(subset=['Job ID'], inplace=True)
            df_job_estimates = df_job_estimates[~df_job_estimates['Job ID'].str.contains('Total', na=False)]
            df_job_estimates = df_job_estimates[~df_job_estimates['Job ID'].astype(str).str.contains('Report', na=False)]
            df_job_estimates = df_job_estimates[df_job_estimates['Phase ID'] != 'Total']
            df_job_estimates.dropna(subset=['Cost Code ID', 'Phase ID'], inplace=True)

            # Clean numeric columns
            df_sales_journal = self.to_numeric_safe(df_sales_journal, ['Billed', 'Amt Recvd'])
            df_job_master = self.to_numeric_safe(df_job_master, ['Contract Amount'])
            df_job_ledger = self.to_numeric_safe(df_job_ledger, ['Debit Amt', 'Credit Amt'])
            df_job_estimates = self.to_numeric_safe(df_job_estimates, ['Est. Expenses'])

            # Clean the job ledger dataframe
            df_job_ledger = self.clean_job_ledger(df_job_ledger)

            self.combine_and_save(df_sales_journal, df_job_master, df_job_ledger, df_job_estimates, parent_window=self.root)

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}", parent=self.root)

    def combine_and_save(self, df_sales_journal, df_job_master, df_job_ledger, df_job_estimates, parent_window=None):
        if parent_window is None:
            parent_window = self.root
        try:
            # --- 1. RENAME & CLEAN DataFrames ---
            for df in [df_sales_journal, df_job_master, df_job_ledger, df_job_estimates]:
                df.columns = [col.strip() for col in df.columns]

            # --- 2. CREATE THE FOUR DATAFRAMES ---
            df_transactions = self._generate_transactions_df(df_job_ledger)
            df_revenue = self._generate_revenue_df(df_sales_journal, df_job_ledger)
            df_expenses = self._generate_expenses_df(df_job_estimates, df_job_ledger)
            df_summary = self._generate_summary_df(df_job_master, df_revenue, df_expenses)

            # --- 3. SAVE TO EXCEL ---
            output_path = self.output_file_path.get()
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df_summary.to_excel(writer, sheet_name='Job Summary', index=False)
                df_revenue.to_excel(writer, sheet_name='Job Revenue', index=False)
                df_expenses.to_excel(writer, sheet_name='Job Expenses', index=False)
                df_transactions.to_excel(writer, sheet_name='Job Transactions', index=False)

            # --- 4. ADD HYPERLINKS AND FORMATTING ---
            wb = load_workbook(output_path)
            self._add_hyperlinks(wb, df_summary, df_expenses, df_revenue, df_transactions)
            self._apply_formatting(wb, df_summary, df_expenses, df_revenue, df_transactions)
            wb.save(output_path)

            # --- 5. TEMPORARY CSV EXPORT ---
            df_summary.to_csv(output_path.replace('.xlsx', '_summary.csv'), index=False)
            df_expenses.to_csv(output_path.replace('.xlsx', '_expenses.csv'), index=False)
            df_revenue.to_csv(output_path.replace('.xlsx', '_revenue.csv'), index=False)
            df_transactions.to_csv(output_path.replace('.xlsx', '_transactions.csv'), index=False)

            messagebox.showinfo("Success", f"Combined file saved to {output_path}")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred during merging: {e}", parent=parent_window)

    def _generate_transactions_df(self, df_job_ledger):
        df_transactions = df_job_ledger.copy()
        df_transactions['Amount'] = df_transactions['Credit Amt'] - df_transactions['Debit Amt']
        df_transactions = df_transactions[['Job ID', 'Cost Code ID', 'Phase Description', 'Phase ID', 'Trx Date', 'Trans Description', 'Amount']]
        df_transactions.rename(columns={'Trx Date': 'Date', 'Trans Description': 'Description'}, inplace=True)
        return df_transactions

    def _generate_revenue_df(self, df_sales_journal, df_job_ledger):
        df = df_sales_journal.copy()
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        return df

    def _generate_expenses_df(self, df_job_estimates, df_job_ledger):
        # Actual Expenses
        actual_expenses = df_job_ledger.groupby(['Job ID', 'Cost Code ID', 'Phase ID'])['Debit Amt'].sum().reset_index()
        actual_expenses.rename(columns={'Debit Amt': 'Actual Expenses'}, inplace=True)

        # Merge with estimates
        df_expenses = pd.merge(df_job_estimates, actual_expenses, on=['Job ID', 'Cost Code ID', 'Phase ID'], how='left')
        df_expenses.rename(columns={'Est. Expenses': 'Est. Expenses'}, inplace=True)
        
        # Percent Complete
        df_expenses['Percent Complete'] = (df_expenses['Actual Expenses'] / df_expenses['Est. Expenses']).fillna(0)
        
        return df_expenses[['Job ID', 'Cost Code ID', 'Phase Description', 'Phase ID', 'Est. Expenses', 'Actual Expenses', 'Percent Complete']]

    def _generate_summary_df(self, df_job_master, df_revenue, df_expenses):
        # Billed and Received
        billed_recvd = df_revenue.groupby('Job ID').agg(
            Amt_Billed=('Billed', 'sum'),
            Amt_Recvd=('Amt Recvd', 'sum')
        ).reset_index()

        # Estimated and Actual Expenses
        expenses = df_expenses.groupby('Job ID').agg(
            Estimated_Expenses=('Est. Expenses', 'sum'),
            Actual_Expenses=('Actual Expenses', 'sum')
        ).reset_index()

        # Percent Complete based on phase_id_criteria
        phase_expenses = df_expenses[df_expenses['Phase ID'] == self.phase_id_criteria]
        phase_expenses_agg = phase_expenses.groupby('Job ID').agg(
            Phase_Est_Expenses=('Est. Expenses', 'sum'),
            Phase_Act_Expenses=('Actual Expenses', 'sum')
        ).reset_index()

        phase_expenses_agg['Percent Complete'] = (phase_expenses_agg['Phase_Act_Expenses'] / phase_expenses_agg['Phase_Est_Expenses']).fillna(0)

        # Merge all
        df_summary = pd.merge(df_job_master, billed_recvd, on='Job ID', how='left')
        df_summary = pd.merge(df_summary, expenses, on='Job ID', how='left')
        df_summary = pd.merge(df_summary, phase_expenses_agg[['Job ID', 'Percent Complete']], on='Job ID', how='left')
        df_summary['Percent Complete'].fillna(0, inplace=True)

        # Rename for consistency
        df_summary.rename(columns={
            'Amt_Billed': 'Amt Billed',
            'Amt_Recvd': 'Amt Recvd',
            'Estimated_Expenses': 'Estimated Expenses',
            'Actual_Expenses': 'Actual Expenses'
        }, inplace=True)

        # Calculations
        df_summary['Left To Bill'] = df_summary['Contract Amount'] - df_summary['Amt Billed']
        df_summary['Left to Receive'] = df_summary['Amt Billed'] - df_summary['Amt Recvd']
        df_summary['Expense Diff'] = df_summary['Estimated Expenses'] - df_summary['Actual Expenses']

        return df_summary

    def _add_hyperlinks(self, wb, df_summary, df_expenses, df_revenue, df_transactions):
        ws_summary = wb['Job Summary']
        ws_expenses = wb['Job Expenses']

        # Hyperlinks from Summary to Revenue
        revenue_link_cols = ['Amt Billed', 'Left To Bill', 'Amt Recvd', 'Left to Receive']
        for col_name in revenue_link_cols:
            if col_name in df_summary.columns:
                col_idx = df_summary.columns.get_loc(col_name) + 1
                for row_idx, job_id in enumerate(df_summary['Job ID'], 2):
                    target_row_idx = df_revenue[df_revenue['Job ID'] == job_id].index.min()
                    if pd.notna(target_row_idx):
                        target_cell = f"A{target_row_idx + 2}"
                        ws_summary.cell(row=row_idx, column=col_idx).hyperlink = f"#'Job Revenue'!{target_cell}"

        # Hyperlinks from Summary to Expenses
        expenses_link_cols = ['Estimated Expenses', 'Actual Expenses', 'Expense Diff', 'Percent Complete']
        for col_name in expenses_link_cols:
            if col_name in df_summary.columns:
                col_idx = df_summary.columns.get_loc(col_name) + 1
                for row_idx, job_id in enumerate(df_summary['Job ID'], 2):
                    target_row_idx = df_expenses[df_expenses['Job ID'] == job_id].index.min()
                    if pd.notna(target_row_idx):
                        target_cell = f"A{target_row_idx + 2}"
                        ws_summary.cell(row=row_idx, column=col_idx).hyperlink = f"#'Job Expenses'!{target_cell}"

        # Hyperlinks from Expenses to Transactions
        if 'Job ID' in df_expenses.columns:
            col_idx = df_expenses.columns.get_loc('Job ID') + 1
            for row_idx, job_id in enumerate(df_expenses['Job ID'], 2):
                target_row_idx = df_transactions[df_transactions['Job ID'] == job_id].index.min()
                if pd.notna(target_row_idx):
                    target_cell = f"A{target_row_idx + 2}"
                    ws_expenses.cell(row=row_idx, column=col_idx).hyperlink = f"#'Job Transactions'!{target_cell}"

    def _apply_formatting(self, wb, df_summary, df_expenses, df_revenue, df_transactions):
        # Define styles
        currency_style = NamedStyle(name='currency', number_format='"$"#,##0.00;[Red]"$"#,##0.00')
        percent_style = NamedStyle(name='percent', number_format='0.00%')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        if 'currency' not in wb.style_names:
            wb.add_named_style(currency_style)
        if 'percent' not in wb.style_names:
            wb.add_named_style(percent_style)

        # Format sheets
        self._format_sheet(wb['Job Summary'], df_summary, ['Contract Amount', 'Amt Billed', 'Left To Bill', 'Amt Recvd', 'Left to Receive', 'Estimated Expenses', 'Actual Expenses', 'Expense Diff'], ['Percent Complete'], None)
        self._format_sheet(wb['Job Expenses'], df_expenses, ['Est. Expenses', 'Actual Expenses'], ['Percent Complete'], {'col': 'Percent Complete', 'base_col': 'Phase ID', 'base_val': self.phase_id_criteria, 'red_fill': red_fill, 'green_fill': green_fill})
        self._format_sheet(wb['Job Revenue'], df_revenue, ['Billed', 'Amt Recvd'], [], None)
        self._format_sheet(wb['Job Transactions'], df_transactions, ['Amount'], [], None)

        # Auto-fit columns
        for sheet_name in wb.sheetnames:
            self._auto_fit_columns(wb[sheet_name])

    def _format_sheet(self, ws, df, currency_cols, percent_cols, conditional_format_info):
        for col_name in currency_cols:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                    for c in cell:
                        c.style = 'currency'
        
        for col_name in percent_cols:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                    for c in cell:
                        c.style = 'percent'

        if conditional_format_info:
            col_name = conditional_format_info['col']
            base_col = conditional_format_info['base_col']
            base_val = conditional_format_info['base_val']
            red_fill = conditional_format_info['red_fill']
            green_fill = conditional_format_info['green_fill']

            base_pc = df[df[base_col] == base_val].groupby('Job ID')[col_name].first()
            col_idx = df.columns.get_loc(col_name) + 1

            for row_idx, job_id in enumerate(df['Job ID'], 2):
                if job_id in base_pc:
                    target_pc = base_pc[job_id]
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        if cell.value > target_pc:
                            cell.fill = red_fill
                        elif cell.value < target_pc:
                            cell.fill = green_fill

    def _auto_fit_columns(self, ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width


if __name__ == "__main__":
    root = tk.Tk()
    app = ProExcelMergerApp(root)
    root.mainloop()
