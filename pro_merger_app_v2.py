import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import pandas as pd
from PIL import Image, ImageTk
import webbrowser
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, NamedStyle
import numpy as np

class ProExcelMergerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Pro Merger v2.0")

        # File paths
        self.sales_journal_path = tk.StringVar()
        self.job_master_path = tk.StringVar()
        self.job_ledger_path = tk.StringVar()
        self.job_estimates_path = tk.StringVar()
        self.output_file_path = tk.StringVar()

        self.config_file = 'merger_config_v2.json'
        self.version = "2.0.4"
        self.tab_order = ['Job Summary', 'Job Revenue', 'Job Expenses', 'Job Transactions', 'Unlinked Items']
        self.tab_configs = {
            'Job Summary': {
                'columns': ['Job ID', 'Project Manager', 'Contract Amount', 'Amt Billed', 'Left To Bill', 'Amt Recvd', 'Left to Receive', 'Estimated Expenses', 'Actual Expenses', 'Expense Diff', 'Percent Complete'],
                'renames': {}
            },
            'Job Revenue': {
                'columns': ['Job ID', 'Billed', 'Amt Recvd'],
                'renames': {}
            },
            'Job Expenses': {
                'columns': ['Job ID', 'Cost Code ID', 'Phase Description', 'Phase ID', 'Est. Expenses', 'Actual Expenses', 'Percent Complete'],
                'renames': {}
            },
            'Job Transactions': {
                'columns': ['Job ID', 'Cost Code ID', 'Phase Description', 'Phase ID', 'Date', 'Description', 'Amount'],
                'renames': {}
            },
            'Unlinked Items': {
                'columns': ['Source File', 'Status', 'Job ID', 'Cost Code ID', 'Phase ID', 'Trans Description', 'Debit Amt', 'Credit Amt', 'Billed', 'Amt Recvd', 'Contract Amount', 'Project Manager'],
                'renames': {}
            }
        }
        self.percent_complete_phase_id = 'Supervision-Wages'

        self.required_input_columns = {
            'sales_journal': {'Job ID', 'Amt Recvd'},
            'job_master': {'Job ID', 'Contract Amount', 'Project Manager'},
            'job_ledger': {'Job ID', 'Cost Code ID', 'Phase ID', 'Trans Description', 'Debit Amt', 'Credit Amt', 'Trx Date'},
            'job_estimates': {'Job ID', 'Cost Code ID', 'Phase ID', 'Est. Expenses'}
        }

        self.load_app_configuration()
        self.create_menu()

        # Main frame
        self.main_frame = tk.Frame(root, padx=10, pady=10)
        self.main_frame.pack(padx=10, pady=10)

        # Add logo
        self.set_logo("C:/Users/dale/OneDrive/Documents/ISI/BnL/Sage/Images/BnL.Logo.jpg", self.main_frame)

        # File selection UI
        row_num = 1
        # Sales Journal
        tk.Label(self.main_frame, text="Custom Sales Journal:").grid(row=row_num, column=0, padx=5, pady=5, sticky='w')
        tk.Entry(self.main_frame, textvariable=self.sales_journal_path, width=50, state='readonly').grid(row=row_num, column=1, padx=5, pady=5)
        tk.Button(self.main_frame, text="Browse...", command=self.browse_sales_journal).grid(row=row_num, column=2, padx=5, pady=5)
        row_num += 1

        # Job Master File
        tk.Label(self.main_frame, text="Custom Job Master File:").grid(row=row_num, column=0, padx=5, pady=5, sticky='w')
        tk.Entry(self.main_frame, textvariable=self.job_master_path, width=50, state='readonly').grid(row=row_num, column=1, padx=5, pady=5)
        tk.Button(self.main_frame, text="Browse...", command=self.browse_job_master).grid(row=row_num, column=2, padx=5, pady=5)
        row_num += 1

        # Job Ledger
        tk.Label(self.main_frame, text="Custom Job Ledger - A:").grid(row=row_num, column=0, padx=5, pady=5, sticky='w')
        tk.Entry(self.main_frame, textvariable=self.job_ledger_path, width=50, state='readonly').grid(row=row_num, column=1, padx=5, pady=5)
        tk.Button(self.main_frame, text="Browse...", command=self.browse_job_ledger).grid(row=row_num, column=2, padx=5, pady=5)
        row_num += 1

        # Job Estimates
        tk.Label(self.main_frame, text="Estimated Job Expenses - A:").grid(row=row_num, column=0, padx=5, pady=5, sticky='w')
        tk.Entry(self.main_frame, textvariable=self.job_estimates_path, width=50, state='readonly').grid(row=row_num, column=1, padx=5, pady=5)
        tk.Button(self.main_frame, text="Browse...", command=self.browse_job_estimates).grid(row=row_num, column=2, padx=5, pady=5)
        row_num += 1

        # Output file path
        tk.Label(self.main_frame, text="Output File:").grid(row=row_num, column=0, padx=5, pady=5, sticky='w')
        tk.Entry(self.main_frame, textvariable=self.output_file_path, width=50, state='readonly').grid(row=row_num, column=1, padx=5, pady=5)
        tk.Button(self.main_frame, text="Browse...", command=self.browse_output_file).grid(row=row_num, column=2, padx=5, pady=5)

        # Go button
        self.go_button = tk.Button(root, text="Go", command=self.go_process, state=tk.DISABLED)
        self.go_button.pack(pady=10)

    def create_menu(self):
        self.menubar = tk.Menu(self.root)
        
        self.filemenu = tk.Menu(self.menubar, tearoff=0)
        self.filemenu.add_command(label="Select Logo", command=self.select_logo_config)
        self.filemenu.add_command(label="Configure Columns", command=self.open_column_config)
        self.filemenu.add_separator()
        self.filemenu.add_command(label="Exit", command=self.root.quit)
        self.menubar.add_cascade(label="File", menu=self.filemenu)

        helpmenu = tk.Menu(self.menubar, tearoff=0)
        helpmenu.add_command(label="About", command=self.show_about)
        self.menubar.add_cascade(label="Help", menu=helpmenu)

        self.root.config(menu=self.menubar)

    def load_app_configuration(self):
        try:
            with open(self.config_file, 'r') as f:
                config = json.load(f)
                self.tab_order = config.get('tab_order', self.tab_order)
                self.tab_configs = config.get('tab_configs', self.tab_configs)
                self.percent_complete_phase_id = config.get('percent_complete_phase_id', self.percent_complete_phase_id)
                loaded_reqs = config.get('required_input_columns', self.required_input_columns)
                self.required_input_columns = {k: set(v) for k, v in loaded_reqs.items()}
        except FileNotFoundError:
            pass # Keep defaults if file not found

    def save_app_configuration(self):
        reqs_to_save = {k: list(v) for k, v in self.required_input_columns.items()}
        config = {
            'tab_order': self.tab_order,
            'tab_configs': self.tab_configs,
            'percent_complete_phase_id': self.percent_complete_phase_id,
            'required_input_columns': reqs_to_save
        }
        with open(self.config_file, 'w') as f:
            json.dump(config, f, indent=4)
        messagebox.showinfo("Saved", "Configuration saved successfully.")

    def open_column_config(self):
        config_window = tk.Toplevel(self.root)
        config_window.title("Configure Report")

        # Tab selection
        tk.Label(config_window, text="Configure Tab:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        selected_tab = tk.StringVar(value=self.tab_order[0])
        tab_menu = tk.OptionMenu(config_window, selected_tab, *self.tab_order)
        tab_menu.grid(row=0, column=1, padx=5, pady=5, sticky='ew')

        # Column listbox
        list_frame = tk.Frame(config_window)
        list_frame.grid(row=1, column=0, columnspan=2, padx=10, pady=10)
        self.col_listbox = tk.Listbox(list_frame, selectmode=tk.SINGLE, width=50, height=15)
        self.col_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = tk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.col_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.col_listbox.config(yscrollcommand=scrollbar.set)

        # Buttons
        btn_frame = tk.Frame(config_window)
        btn_frame.grid(row=1, column=2, padx=5, pady=5)
        tk.Button(btn_frame, text="Up", command=lambda: self.move_col_up(selected_tab.get())).pack(fill=tk.X)
        tk.Button(btn_frame, text="Down", command=lambda: self.move_col_down(selected_tab.get())).pack(fill=tk.X)
        tk.Button(btn_frame, text="Rename", command=lambda: self.rename_col(selected_tab.get(), config_window)).pack(fill=tk.X)
        tk.Button(btn_frame, text="Remove", command=lambda: self.remove_col(selected_tab.get())).pack(fill=tk.X)

        # --- Phase ID selection ---
        phase_id_frame = tk.Frame(config_window)
        phase_id_frame.grid(row=2, column=0, columnspan=3, pady=10, sticky='ew')
        tk.Label(phase_id_frame, text="Select Phase ID for Percent Complete calculation:").pack(side=tk.LEFT, padx=5)

        phase_ids = ['Supervision-Wages'] # Default
        if self.job_estimates_path.get():
            try:
                df = pd.read_excel(self.job_estimates_path.get())
                if 'Phase ID' in df.columns:
                    phase_ids = sorted(df['Phase ID'].dropna().unique().tolist())
            except Exception as e:
                messagebox.showerror("Error", f"Could not load Phase IDs: {e}", parent=config_window)

        selected_phase_id = tk.StringVar(value=self.percent_complete_phase_id)
        phase_id_menu = tk.OptionMenu(phase_id_frame, selected_phase_id, *phase_ids)
        phase_id_menu.pack(side=tk.LEFT, padx=5)

        def update_phase_id(*args):
            self.percent_complete_phase_id = selected_phase_id.get()

        selected_phase_id.trace('w', update_phase_id)

        # Save button
        tk.Button(config_window, text="Save Configuration", command=self.save_app_configuration).grid(row=3, column=0, columnspan=3, pady=10)

        def update_listbox_for_tab_wrapper(*args):
            self.update_listbox_for_tab(selected_tab.get())

        selected_tab.trace('w', update_listbox_for_tab_wrapper)
        self.update_listbox_for_tab(selected_tab.get()) # Initial population

    def update_listbox_for_tab(self, tab):
        self.col_listbox.delete(0, tk.END)
        if tab in self.tab_configs:
            for col in self.tab_configs[tab]['columns']:
                display_name = self.tab_configs[tab]['renames'].get(col, col)
                self.col_listbox.insert(tk.END, display_name)

    def move_col_up(self, tab):
        try:
            idx = self.col_listbox.curselection()[0]
            if idx > 0:
                cols = self.tab_configs[tab]['columns']
                cols.insert(idx - 1, cols.pop(idx))
                self.update_listbox_for_tab(tab)
                self.col_listbox.selection_set(idx - 1)
        except IndexError:
            pass

    def move_col_down(self, tab):
        try:
            idx = self.col_listbox.curselection()[0]
            cols = self.tab_configs[tab]['columns']
            if idx < len(cols) - 1:
                cols.insert(idx + 1, cols.pop(idx))
                self.update_listbox_for_tab(tab)
                self.col_listbox.selection_set(idx + 1)
        except IndexError:
            pass

    def rename_col(self, tab, parent_window):
        try:
            idx = self.col_listbox.curselection()[0]
            original_col_name = self.tab_configs[tab]['columns'][idx]
            old_display_name = self.tab_configs[tab]['renames'].get(original_col_name, original_col_name)
            
            new_name = simpledialog.askstring("Rename Column", f"Enter new name for '{old_display_name}':", parent=parent_window)
            if new_name:
                self.tab_configs[tab]['renames'][original_col_name] = new_name
                self.update_listbox_for_tab(tab)
                self.col_listbox.selection_set(idx)
        except IndexError:
            messagebox.showwarning("Warning", "Please select a column to rename.", parent=parent_window)

    def remove_col(self, tab):
        try:
            idx = self.col_listbox.curselection()[0]
            self.tab_configs[tab]['columns'].pop(idx)
            self.update_listbox_for_tab(tab)
        except IndexError:
            messagebox.showwarning("Warning", "Please select a column to remove.")


    def show_about(self):
        about_window = tk.Toplevel(self.root)
        about_window.title("About Pro Excel Merger")
        
        logo_path = "C:/Users/dale/OneDrive/Documents/ISI/BnL/Sage/Images/IDS.Logo.png"
        logo_image = Image.open(logo_path)
        logo_image = logo_image.resize((100, 100), Image.LANCZOS)
        self.about_logo_photo = ImageTk.PhotoImage(logo_image)
        
        tk.Label(about_window, image=self.about_logo_photo).pack(pady=10)
        tk.Label(about_window, text=f"Version: {self.version}").pack(pady=5)
        
        tk.Label(about_window, text="Contact: Dale Linn").pack(pady=5)
        email = tk.Label(about_window, text="richard.linn@biz-solve.com", fg="blue", cursor="hand2")
        email.pack()
        email.bind("<Button-1>", lambda e: webbrowser.open_new("mailto:richard.linn@biz-solve.com"))

        tk.Label(about_window, text="For new versions, please visit:").pack(pady=5)
        link = tk.Label(about_window, text="https://github.com/dlerhetal/BnL.ExcelMerge/releases", fg="blue", cursor="hand2")
        link.pack()
        link.bind("<Button-1>", lambda e: webbrowser.open_new("https://github.com/dlerhetal/BnL.ExcelMerge/releases"))

    def set_logo(self, logo_path, frame):
        try:
            logo_image = Image.open(logo_path)
            aspect_ratio = logo_image.width / logo_image.height
            logo_image = logo_image.resize((int(50 * aspect_ratio), 50), Image.LANCZOS)
            self.logo_photo = ImageTk.PhotoImage(logo_image)
            
            if hasattr(self, 'logo_label'):
                self.logo_label.config(image=self.logo_photo)
            else:
                self.logo_label = tk.Label(frame, image=self.logo_photo)
                self.logo_label.grid(row=0, column=0, columnspan=3, pady=10)

        except Exception as e:
            messagebox.showerror("Error", f"Could not load logo: {e}")

    def select_logo_config(self):
        filename = filedialog.askopenfilename(filetypes=[
            ("Image files", "*.jpg *.jpeg *.png *.gif *.bmp"),
            ("All files", "*.*",)
        ])
        if filename:
            self.set_logo(filename, self.main_frame)

    def validate_input_file(self, filename, file_type):
        try:
            df_cols = set(pd.read_excel(filename, nrows=0).columns)

            required = self.required_input_columns[file_type]
            missing_cols = required - df_cols

            is_sales_journal = (file_type == 'sales_journal')
            has_billed_or_amount = ('Billed' in df_cols or 'Amount' in df_cols)

            has_missing = bool(missing_cols)
            if is_sales_journal and not has_billed_or_amount:
                has_missing = True

            if has_missing:
                required_display = sorted(list(required))
                if is_sales_journal:
                    required_display.append('Billed (or Amount)')
                    required_display = sorted(required_display)

                title = "Oops! Column Mismatch"
                message = (
                    "OOPS! THAT FILE DOES NOT MATCH YOUR COLUMN CONFIGURATION THAT YOU SET. "
                    "DID YOU MEAN TO DO THAT?\n\n"
                    "HERE'S WHAT I HAVE:\n"
                    f"{', '.join(required_display)}"
                )
                
                return messagebox.askyesno(title, message, parent=self.root)

            return True
        except Exception as e:
            messagebox.showerror("Error Reading File", f"Could not read the columns from the file: {e}")
            return False

    def browse_sales_journal(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if filename:
            if self.validate_input_file(filename, 'sales_journal'):
                self.sales_journal_path.set(filename)
            else:
                self.sales_journal_path.set('')
            self.check_inputs()

    def browse_job_master(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if filename:
            if self.validate_input_file(filename, 'job_master'):
                self.job_master_path.set(filename)
            else:
                self.job_master_path.set('')
            self.check_inputs()

    def browse_job_ledger(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if filename:
            if self.validate_input_file(filename, 'job_ledger'):
                self.job_ledger_path.set(filename)
            else:
                self.job_ledger_path.set('')
            self.check_inputs()

    def browse_job_estimates(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if filename:
            if self.validate_input_file(filename, 'job_estimates'):
                self.job_estimates_path.set(filename)
            else:
                self.job_estimates_path.set('')
            self.check_inputs()

    def browse_output_file(self):
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if filename:
            self.output_file_path.set(filename)
            self.check_inputs()

    def check_inputs(self, *args):
        if (self.sales_journal_path.get() and
            self.job_master_path.get() and
            self.job_ledger_path.get() and
            self.job_estimates_path.get() and
            self.output_file_path.get()):
            self.go_button.config(state=tk.NORMAL)
            # self.filemenu.entryconfig("Configure Columns", state=tk.NORMAL) # Disabled for now
        else:
            self.go_button.config(state=tk.DISABLED)
            # if hasattr(self, 'filemenu'):
                # self.filemenu.entryconfig("Configure Columns", state=tk.DISABLED)

    def clean_job_ledger(self, df_in):
        df = df_in.copy()
        df.replace('', np.nan, inplace=True)
        
        # Forward fill missing values in specified columns
        cols_to_ffill = ['Job ID', 'Cost Code ID', 'Phase Description', 'Phase ID']
        for col in cols_to_ffill:
            if col in df.columns:
                df[col] = df[col].ffill()

        # Drop rows where 'Trans Description' is NaN, which are likely empty or summary rows
        df.dropna(subset=['Trans Description'], inplace=True)

        # Remove rows containing 'Total' in 'Job ID' or 'Trans Description'
        df = df[~df['Job ID'].astype(str).str.contains('Total', na=False)]
        df = df[~df['Job ID'].astype(str).str.contains('Report', na=False)]
        df = df[~df['Trans Description'].astype(str).str.contains('Total', na=False)]
        df = df[df['Phase ID'] != 'Total']

        return df

    def to_numeric_safe(self, df, columns):
        for col in columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col].astype(str).str.replace('[$,]', '', regex=True), errors='coerce').fillna(0)
        return df

    def go_process(self):
        try:
            # --- 1. Load Raw Data and Add Fingerprints ---
            unlinked_items = []
            
            df_sales_journal_raw = pd.read_excel(self.sales_journal_path.get()).reset_index().rename(columns={'index': 'original_index'})
            df_job_master_raw = pd.read_excel(self.job_master_path.get()).reset_index().rename(columns={'index': 'original_index'})
            df_job_ledger_raw = pd.read_excel(self.job_ledger_path.get()).reset_index().rename(columns={'index': 'original_index'})
            df_job_estimates_raw = pd.read_excel(self.job_estimates_path.get()).reset_index().rename(columns={'index': 'original_index'})

            # --- 2. Create working copies ---
            df_sales_journal = df_sales_journal_raw.copy()
            df_job_master = df_job_master_raw.copy()
            df_job_ledger = df_job_ledger_raw.copy()
            df_job_estimates = df_job_estimates_raw.copy()

            # --- 3. Perform Initial Cleaning and Capture Dropped Rows ---
            
            # Clean Sales Journal
            df_sales_journal.dropna(subset=['Job ID'], inplace=True)

            # Clean Job Master
            if 'Unnamed: 0' in df_job_master.columns and 'Job ID' not in df_job_master.columns:
                df_job_master.rename(columns={'Unnamed: 0': 'Job ID'}, inplace=True)

            # Clean Job Estimates
            est_pre_clean_indices = df_job_estimates['original_index']
            df_job_estimates.dropna(subset=['Job ID'], inplace=True)
            df_job_estimates = df_job_estimates[~df_job_estimates['Job ID'].str.contains('Total', na=False)]
            df_job_estimates = df_job_estimates[~df_job_estimates['Job ID'].str.contains('Report', na=False)]
            df_job_estimates = df_job_estimates[df_job_estimates['Phase ID'] != 'Total']
            df_job_estimates.dropna(subset=['Cost Code ID', 'Phase ID'], inplace=True)
            est_post_clean_indices = df_job_estimates['original_index']
            dropped_est_indices = est_pre_clean_indices[~est_pre_clean_indices.isin(est_post_clean_indices)]
            if not dropped_est_indices.empty:
                dropped_est_df = df_job_estimates_raw[df_job_estimates_raw['original_index'].isin(dropped_est_indices)].copy()
                dropped_est_df['Source File'] = 'Job Estimates'
                dropped_est_df['Status'] = 'Dropped during initial cleaning'
                unlinked_items.append(dropped_est_df)

            # Clean Job Ledger
            ledger_pre_clean_indices = df_job_ledger['original_index']
            df_job_ledger = self.clean_job_ledger(df_job_ledger)
            ledger_post_clean_indices = df_job_ledger['original_index']
            dropped_ledger_indices = ledger_pre_clean_indices[~ledger_pre_clean_indices.isin(ledger_post_clean_indices)]
            if not dropped_ledger_indices.empty:
                dropped_ledger_df = df_job_ledger_raw[df_job_ledger_raw['original_index'].isin(dropped_ledger_indices)].copy()
                dropped_ledger_df['Source File'] = 'Job Ledger'
                dropped_ledger_df['Status'] = 'Dropped during ledger cleaning (e.g., no description)'
                unlinked_items.append(dropped_ledger_df)

            # --- 4. Clean Numeric Columns (on cleaned data) ---
            df_sales_journal = self.to_numeric_safe(df_sales_journal, ['Billed', 'Amt Recvd'])
            df_job_master = self.to_numeric_safe(df_job_master, ['Contract Amount'])
            df_job_ledger = self.to_numeric_safe(df_job_ledger, ['Debit Amt', 'Credit Amt'])
            df_job_estimates = self.to_numeric_safe(df_job_estimates, ['Est. Expenses'])

            # --- 5. Pass all data to combine and save ---
            self.combine_and_save(
                df_sales_journal, df_job_master, df_job_ledger, df_job_estimates, 
                unlinked_items, parent_window=self.root
            )

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}", parent=self.root)

    def combine_and_save(self, df_sales_journal, df_job_master, df_job_ledger, df_job_estimates, unlinked_items, parent_window=None):
        if parent_window is None:
            parent_window = self.root
        try:
            # --- 1. Generate Base DataFrames and Collect Unlinked Items ---
            df_revenue = self._generate_revenue_df(df_sales_journal, df_job_ledger)
            df_expenses, unlinked_expenses = self._generate_expenses_df(df_job_estimates, df_job_ledger)
            df_summary, unlinked_summary = self._generate_summary_df(df_job_master, df_revenue, df_expenses)

            if not unlinked_expenses.empty:
                unlinked_items.append(unlinked_expenses)
            if not unlinked_summary.empty:
                unlinked_items.append(unlinked_summary)

            base_dfs = {
                'Job Summary': df_summary,
                'Job Revenue': df_revenue,
                'Job Expenses': df_expenses,
                'Job Transactions': self._generate_transactions_df(df_job_ledger)
            }

            # --- 2. Consolidate and Filter Unlinked Items ---
            if unlinked_items:
                all_unlinked_df = pd.concat(unlinked_items, ignore_index=True)
                mask = (
                    all_unlinked_df['Job ID'].notna() &
                    all_unlinked_df['Trans Description'].notna() &
                    (all_unlinked_df['Debit Amt'].notna() | all_unlinked_df['Credit Amt'].notna())
                )
                all_unlinked_df = all_unlinked_df[mask]

                cols_to_drop = ['original_index', '_merge', 'merge_rev', 'merge_exp']
                for col in cols_to_drop:
                    if col in all_unlinked_df.columns:
                        all_unlinked_df.drop(columns=col, inplace=True)
                base_dfs['Unlinked Items'] = all_unlinked_df

            # --- 3. Configure DataFrames ---
            configured_dfs = {}
            for tab_name in self.tab_order:
                if tab_name in self.tab_configs and tab_name in base_dfs:
                    config = self.tab_configs[tab_name]
                    df = base_dfs[tab_name]
                    for col in config['columns']:
                        if col not in df.columns:
                            df[col] = np.nan
                    existing_cols = [col for col in config['columns'] if col in df.columns]
                    output_df = df[existing_cols]
                    output_df = output_df.rename(columns=config['renames'])
                    configured_dfs[tab_name] = output_df

            # --- 4. Save to CSV for debugging ---
            for tab_name, df in configured_dfs.items():
                df.to_csv(f'{tab_name.replace(" ", "_").lower()}_debug.csv', index=False)

            # --- 5. Save to Excel (Main Tabs) ---
            output_path = self.output_file_path.get()
            unlinked_df_to_write = configured_dfs.pop('Unlinked Items', None)

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for tab_name, df in configured_dfs.items():
                    df.to_excel(writer, sheet_name=tab_name, index=False)

            # --- 6. Append Unlinked Sheet and Apply Final Formatting ---
            try:
                wb = load_workbook(output_path)
                
                if unlinked_df_to_write is not None and not unlinked_df_to_write.empty:
                    ws = wb.create_sheet('Unlinked Items')
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    for r in dataframe_to_rows(unlinked_df_to_write, index=False, header=True):
                        ws.append(r)
                
                self._add_hyperlinks(wb, configured_dfs.get('Job Summary'), configured_dfs.get('Job Expenses'), configured_dfs.get('Job Revenue'), configured_dfs.get('Job Transactions'))
                
                if unlinked_df_to_write is not None:
                    configured_dfs['Unlinked Items'] = unlinked_df_to_write

                self._apply_formatting(wb, configured_dfs)
                
                wb.save(output_path)
                messagebox.showinfo("Success", f"Combined file saved to {output_path}")

            except Exception as e:
                messagebox.showerror("Error", f"An error occurred during final Excel processing: {e}", parent=parent_window)

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred during merging: {e}", parent=parent_window)

    def _generate_transactions_df(self, df_job_ledger):
        df_transactions = df_job_ledger.copy()
        df_transactions['Amount'] = df_transactions['Credit Amt'] - df_transactions['Debit Amt']
        df_transactions = df_transactions[['Job ID', 'Cost Code ID', 'Phase Description', 'Phase ID', 'Trx Date', 'Trans Description', 'Amount']]
        df_transactions.rename(columns={'Trx Date': 'Date', 'Trans Description': 'Description'}, inplace=True)
        df_transactions['Date'] = pd.to_datetime(df_transactions['Date']).dt.date
        df_transactions.sort_values(by=['Job ID', 'Date'], inplace=True)
        df_transactions.reset_index(drop=True, inplace=True)
        return df_transactions

    def _generate_revenue_df(self, df_sales_journal, df_job_ledger):
        df = df_sales_journal.copy()
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        if 'Amount' in df.columns and 'Billed' not in df.columns:
            df.rename(columns={'Amount': 'Billed'}, inplace=True)
        df.sort_values(by='Job ID', inplace=True)
        df.reset_index(drop=True, inplace=True)
        return df

    def _generate_expenses_df(self, df_job_estimates, df_job_ledger):
        # Actual Expenses
        actual_expenses = df_job_ledger.groupby(['Job ID', 'Cost Code ID', 'Phase ID'])['Debit Amt'].sum().reset_index()
        actual_expenses.rename(columns={'Debit Amt': 'Actual Expenses'}, inplace=True)

        # Merge with estimates using an outer join to find unlinked items
        df_merged = pd.merge(df_job_estimates, actual_expenses, on=['Job ID', 'Cost Code ID', 'Phase ID'], how='outer', indicator=True)
        
        # Separate the unlinked items (from actuals but not in estimates)
        unlinked_expenses_df = df_merged[df_merged['_merge'] == 'right_only'].copy()
        unlinked_expenses_df['Source File'] = 'Job Ledger'
        unlinked_expenses_df['Status'] = 'Actual expense with no matching estimate'

        # Create the main expenses df (everything that was in estimates)
        df_expenses = df_merged[df_merged['_merge'] != 'right_only'].copy()
        df_expenses.rename(columns={'Est. Expenses': 'Est. Expenses'}, inplace=True)
        
        # Percent Complete
        df_expenses['Percent Complete'] = (df_expenses['Actual Expenses'] / df_expenses['Est. Expenses']).fillna(0)
        df_expenses.sort_values(by=['Job ID', 'Cost Code ID', 'Phase ID'], inplace=True)
        df_expenses.reset_index(drop=True, inplace=True)
        
        final_cols = ['Job ID', 'Cost Code ID', 'Phase Description', 'Phase ID', 'Est. Expenses', 'Actual Expenses', 'Percent Complete']
        return df_expenses[final_cols], unlinked_expenses_df

    def _generate_summary_df(self, df_job_master, df_revenue, df_expenses):
        unlinked_items = []
        # Billed and Received
        billed_recvd = df_revenue.groupby('Job ID').agg(
            Amt_Billed=('Billed', 'sum'),
            Amt_Recvd=('Amt Recvd', 'sum')
        ).reset_index()

        # Estimated and Actual Expenses
        expenses = df_expenses.groupby('Job ID').agg(
            Estimated_Expenses=('Est. Expenses', 'sum'),
            Actual_Expenses=('Actual Expenses', 'sum')
        ).reset_index()

        # Percent Complete based on phase_id_criteria
        phase_expenses = df_expenses[df_expenses['Phase ID'] == self.percent_complete_phase_id]
        phase_expenses_agg = phase_expenses.groupby('Job ID').agg(
            Phase_Est_Expenses=('Est. Expenses', 'sum'),
            Phase_Act_Expenses=('Actual Expenses', 'sum')
        ).reset_index()

        phase_expenses_agg['Percent Complete'] = (phase_expenses_agg['Phase_Act_Expenses'] / phase_expenses_agg['Phase_Est_Expenses']).fillna(0)

        # --- Merge all, finding orphans ---
        df_summary = pd.merge(df_job_master, billed_recvd, on='Job ID', how='outer', indicator='merge_rev')
        unlinked_revenue = df_summary[df_summary['merge_rev'] == 'right_only'].copy()
        if not unlinked_revenue.empty:
            unlinked_revenue['Source File'] = 'Sales Journal'
            unlinked_revenue['Status'] = 'Revenue item with no matching Job Master record'
            unlinked_items.append(unlinked_revenue)

        df_summary = df_summary[df_summary['merge_rev'] != 'right_only'] # Continue with linked items

        df_summary = pd.merge(df_summary, expenses, on='Job ID', how='outer', indicator='merge_exp')
        unlinked_expenses = df_summary[df_summary['merge_exp'] == 'right_only'].copy()
        if not unlinked_expenses.empty:
            unlinked_expenses['Source File'] = 'Job Expenses Tab'
            unlinked_expenses['Status'] = 'Aggregated expense with no matching Job Master record'
            unlinked_items.append(unlinked_expenses)
        
        df_summary = df_summary[df_summary['merge_exp'] != 'right_only'] # Continue with linked items

        df_summary = pd.merge(df_summary, phase_expenses_agg[['Job ID', 'Percent Complete']], on='Job ID', how='left')
        df_summary['Percent Complete'] = df_summary['Percent Complete'].fillna(0)

        # Rename for consistency
        df_summary.rename(columns={
            'Amt_Billed': 'Amt Billed',
            'Amt_Recvd': 'Amt Recvd',
            'Estimated_Expenses': 'Estimated Expenses',
            'Actual_Expenses': 'Actual Expenses'
        }, inplace=True)

        # Calculations
        df_summary['Left To Bill'] = df_summary['Contract Amount'] - df_summary['Amt Billed'].abs()
        df_summary.loc[df_summary['Contract Amount'] == 0, 'Left To Bill'] = pd.NA
        df_summary['Left to Receive'] = df_summary['Amt Billed'].abs() - df_summary['Amt Recvd']
        df_summary['Expense Diff'] = df_summary['Estimated Expenses'] - df_summary['Actual Expenses']

        unlinked_summary_df = pd.concat(unlinked_items, ignore_index=True) if unlinked_items else pd.DataFrame()

        return df_summary, unlinked_summary_df

    def _add_hyperlinks(self, wb, df_summary, df_expenses, df_revenue, df_transactions):
        ws_summary = wb['Job Summary']
        ws_expenses = wb['Job Expenses']

        # Hyperlinks from Summary to Revenue
        revenue_link_cols = ['Amt Billed', 'Left To Bill', 'Amt Recvd', 'Left to Receive']
        for col_name in revenue_link_cols:
            if col_name in df_summary.columns:
                col_idx = df_summary.columns.get_loc(col_name) + 1
                for row_idx, job_id in enumerate(df_summary['Job ID'], 2):
                    if pd.notna(df_summary.at[row_idx - 2, col_name]):
                        target_row_idx = df_revenue[df_revenue['Job ID'] == job_id].index.min()
                        if pd.notna(target_row_idx):
                            target_cell = f"A{target_row_idx + 2}"
                            ws_summary.cell(row=row_idx, column=col_idx).hyperlink = f"#'Job Revenue'!{target_cell}"

        # Hyperlinks from Summary to Expenses
        expenses_link_cols = ['Estimated Expenses', 'Actual Expenses', 'Expense Diff', 'Percent Complete']
        for col_name in expenses_link_cols:
            if col_name in df_summary.columns:
                col_idx = df_summary.columns.get_loc(col_name) + 1
                for row_idx, job_id in enumerate(df_summary['Job ID'], 2):
                    if pd.notna(df_summary.at[row_idx - 2, col_name]):
                        target_row_idx = df_expenses[df_expenses['Job ID'] == job_id].index.min()
                        if pd.notna(target_row_idx):
                            target_cell = f"A{target_row_idx + 2}"
                            ws_summary.cell(row=row_idx, column=col_idx).hyperlink = f"#'Job Expenses'!{target_cell}"

        # Hyperlinks from Expenses to Transactions
        if 'Job ID' in df_expenses.columns:
            col_idx = df_expenses.columns.get_loc('Job ID') + 1
            for row_idx, job_id in enumerate(df_expenses['Job ID'], 2):
                target_row_idx = df_transactions[df_transactions['Job ID'] == job_id].index.min()
                if pd.notna(target_row_idx):
                    target_cell = f"A{target_row_idx + 2}"
                    ws_expenses.cell(row=row_idx, column=col_idx).hyperlink = f"#'Job Transactions'!{target_cell}"

    def _apply_formatting(self, wb, configured_dfs):
        # Define styles
        currency_style = NamedStyle(name='currency', number_format='"$"#,##0.00;[Red]"$"#,##0.00')
        percent_style = NamedStyle(name='percent', number_format='0.00%')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        if 'currency' not in wb.style_names:
            wb.add_named_style(currency_style)
        if 'percent' not in wb.style_names:
            wb.add_named_style(percent_style)

        # Format sheets
        for tab_name, df in configured_dfs.items():
            ws = wb[tab_name]
            config = self.tab_configs[tab_name]
            renames = config['renames']

            if tab_name == 'Job Summary':
                currency_cols = ['Contract Amount', 'Amt Billed', 'Left To Bill', 'Amt Recvd', 'Left to Receive', 'Estimated Expenses', 'Actual Expenses', 'Expense Diff']
                percent_cols = ['Percent Complete']
                self._format_sheet(ws, df, currency_cols, percent_cols, None, renames)
            elif tab_name == 'Job Expenses':
                currency_cols = ['Est. Expenses', 'Actual Expenses']
                percent_cols = ['Percent Complete']
                conditional_format_info = {'col': 'Percent Complete', 'base_col': 'Phase ID', 'base_val': self.percent_complete_phase_id, 'red_fill': red_fill, 'green_fill': green_fill}
                self._format_sheet(ws, df, currency_cols, percent_cols, conditional_format_info, renames)
            elif tab_name == 'Job Revenue':
                currency_cols = ['Billed', 'Amt Recvd']
                self._format_sheet(ws, df, currency_cols, [], None, renames)
            elif tab_name == 'Job Transactions':
                currency_cols = ['Amount']
                self._format_sheet(ws, df, currency_cols, [], None, renames)

        # Auto-fit columns
        for sheet_name in wb.sheetnames:
            self._auto_fit_columns(wb[sheet_name])

    def _format_sheet(self, ws, df, currency_cols, percent_cols, conditional_format_info, renames):
        for col_name in currency_cols:
            new_col_name = renames.get(col_name, col_name)
            if new_col_name in df.columns:
                col_idx = df.columns.get_loc(new_col_name) + 1
                for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                    for c in cell:
                        c.style = 'currency'
        
        for col_name in percent_cols:
            new_col_name = renames.get(col_name, col_name)
            if new_col_name in df.columns:
                col_idx = df.columns.get_loc(new_col_name) + 1
                for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                    for c in cell:
                        c.style = 'percent'

        if conditional_format_info:
            col_name = conditional_format_info['col']
            new_col_name = renames.get(col_name, col_name)
            base_col = conditional_format_info['base_col']
            new_base_col = renames.get(base_col, base_col)
            base_val = conditional_format_info['base_val']
            red_fill = conditional_format_info['red_fill']
            green_fill = conditional_format_info['green_fill']

            if new_base_col in df.columns and new_col_name in df.columns:
                base_pc = df[df[new_base_col] == base_val].groupby('Job ID')[new_col_name].first()
                col_idx = df.columns.get_loc(new_col_name) + 1

                for row_idx, job_id in enumerate(df['Job ID'], 2):
                    if job_id in base_pc:
                        target_pc = base_pc[job_id]
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            if cell.value > target_pc:
                                cell.fill = red_fill
                            elif cell.value < target_pc:
                                cell.fill = green_fill

    def _auto_fit_columns(self, ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width


if __name__ == "__main__":
    root = tk.Tk()
    app = ProExcelMergerApp(root)
    root.mainloop()
