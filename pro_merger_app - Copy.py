import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import pandas as pd
from PIL import Image, ImageTk
import webbrowser
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

class ProExcelMergerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Pro Excel Merger")

        self.expense_file_path = tk.StringVar()
        self.revenue_file_path = tk.StringVar()
        self.output_file_path = tk.StringVar()
        self.key_columns = ['Job ID', 'Cost Code ID', 'Phase ID']
        self.required_column = None
        self.original_columns = []
        self.rename_map = {}
        self.supervisor_columns = {}
        self.phase_id_criteria = tk.StringVar()
        self.config_file = 'merger_config.json'
        self.version = "1.1.1"

        self.create_menu()

        # Main frame
        self.main_frame = tk.Frame(root, padx=10, pady=10)
        self.main_frame.pack(padx=10, pady=10)

        # Add logo
        self.set_logo("C:/Users/dale/OneDrive/Documents/ISI/BnL/Sage/Images/BnL.Logo.jpg", self.main_frame)

        # Expense file selection
        tk.Label(self.main_frame, text="Expense File:").grid(row=1, column=0, padx=5, pady=5, sticky='w')
        tk.Entry(self.main_frame, textvariable=self.expense_file_path, width=50, state='readonly').grid(row=1, column=1, padx=5, pady=5)
        tk.Button(self.main_frame, text="Browse...", command=self.browse_expense_file).grid(row=1, column=2, padx=5, pady=5)

        # Revenue file selection
        tk.Label(self.main_frame, text="Revenue File:").grid(row=2, column=0, padx=5, pady=5, sticky='w')
        tk.Entry(self.main_frame, textvariable=self.revenue_file_path, width=50, state='readonly').grid(row=2, column=1, padx=5, pady=5)
        tk.Button(self.main_frame, text="Browse...", command=self.browse_revenue_file).grid(row=2, column=2, padx=5, pady=5)

        # Output file path
        tk.Label(self.main_frame, text="Output File:").grid(row=3, column=0, padx=5, pady=5, sticky='w')
        tk.Entry(self.main_frame, textvariable=self.output_file_path, width=50, state='readonly').grid(row=3, column=1, padx=5, pady=5)
        tk.Button(self.main_frame, text="Browse...", command=self.browse_output_file).grid(row=3, column=2, padx=5, pady=5)

        # Next button
        self.next_button = tk.Button(root, text="Configure Columns", command=self.open_column_config, state=tk.DISABLED)
        self.next_button.pack(pady=10)

    def create_menu(self):
        menubar = tk.Menu(self.root)
        
        filemenu = tk.Menu(menubar, tearoff=0)
        filemenu.add_command(label="Select Logo", command=self.select_logo_config)
        filemenu.add_separator()
        filemenu.add_command(label="Exit", command=self.root.quit)
        menubar.add_cascade(label="File", menu=filemenu)

        helpmenu = tk.Menu(menubar, tearoff=0)
        helpmenu.add_command(label="About", command=self.show_about)
        menubar.add_cascade(label="Help", menu=helpmenu)

        self.root.config(menu=menubar)

    def show_about(self):
        about_window = tk.Toplevel(self.root)
        about_window.title("About Pro Excel Merger")
        
        logo_path = "C:/Users/dale/OneDrive/Documents/ISI/BnL/Sage/Images/IDS.Logo.png"
        logo_image = Image.open(logo_path)
        logo_image = logo_image.resize((100, 100), Image.LANCZOS)
        self.about_logo_photo = ImageTk.PhotoImage(logo_image)
        
        tk.Label(about_window, image=self.about_logo_photo).pack(pady=10)
        tk.Label(about_window, text=f"Version: {self.version}").pack(pady=5)
        
        tk.Label(about_window, text="Contact: Dale Linn").pack(pady=5)
        email = tk.Label(about_window, text="richard.linn@biz-solve.com", fg="blue", cursor="hand2")
        email.pack()
        email.bind("<Button-1>", lambda e: webbrowser.open_new("mailto:richard.linn@biz-solve.com"))

        tk.Label(about_window, text="For new versions, please visit:").pack(pady=5)
        link = tk.Label(about_window, text="https://github.com/dlerhetal/BnL.ExcelMerge/releases", fg="blue", cursor="hand2")
        link.pack()
        link.bind("<Button-1>", lambda e: webbrowser.open_new("https://github.com/dlerhetal/BnL.ExcelMerge/releases"))

    def set_logo(self, logo_path, frame):
        try:
            logo_image = Image.open(logo_path)
            aspect_ratio = logo_image.width / logo_image.height
            logo_image = logo_image.resize((int(50 * aspect_ratio), 50), Image.LANCZOS)
            self.logo_photo = ImageTk.PhotoImage(logo_image)
            
            if hasattr(self, 'logo_label'):
                self.logo_label.config(image=self.logo_photo)
            else:
                self.logo_label = tk.Label(frame, image=self.logo_photo)
                self.logo_label.grid(row=0, column=0, columnspan=3, pady=10)

        except Exception as e:
            messagebox.showerror("Error", f"Could not load logo: {e}")

    def select_logo_config(self):
        filename = filedialog.askopenfilename(filetypes=[
            ("Image files", "*.jpg *.jpeg *.png *.gif *.bmp"),
            ("All files", "*.*",)
        ])
        if filename:
            self.set_logo(filename, self.main_frame)

    def browse_expense_file(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if filename:
            self.expense_file_path.set(filename)
            self.check_inputs()

    def browse_revenue_file(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if filename:
            self.revenue_file_path.set(filename)
            self.check_inputs()

    def browse_output_file(self):
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if filename:
            self.output_file_path.set(filename)
            self.check_inputs()

    def check_inputs(self, *args):
        if self.expense_file_path.get() and self.revenue_file_path.get() and self.output_file_path.get():
            self.next_button.config(state=tk.NORMAL)
        else:
            self.next_button.config(state=tk.DISABLED)

    def open_column_config(self):
        try:
            self.df_expense = pd.read_excel(self.expense_file_path.get())
            self.df_revenue = pd.read_excel(self.revenue_file_path.get())

            self.col_config_window = tk.Toplevel(self.root)
            self.col_config_window.title("Configure Columns")

            list_frame = tk.Frame(self.col_config_window)
            list_frame.grid(row=0, column=0, padx=10, pady=10, rowspan=5)

            self.listbox = tk.Listbox(list_frame, selectmode=tk.SINGLE, width=40, height=15)
            self.key_columns_label = tk.Label(self.col_config_window, text=f"Keys: {self.key_columns}")
            self.key_columns_label.grid(row=5, column=0, padx=5, pady=5, sticky='w')

            self.required_column_label = tk.Label(self.col_config_window, text=f"Required Column: {self.required_column if self.required_column else 'None'}")
            self.required_column_label.grid(row=7, column=0, padx=5, pady=5, sticky='w')

            self.load_app_configuration()

            required_col_button = tk.Button(self.col_config_window, text="Select Required Column", command=self.select_required_column)
            required_col_button.grid(row=8, column=0, padx=5, pady=5, sticky='ew')

            reset_required_col_button = tk.Button(self.col_config_window, text="Reset Required Column", command=self.reset_required_column)
            reset_required_col_button.grid(row=8, column=1, padx=5, pady=5, sticky='ew')

            add_percent_complete_button = tk.Button(self.col_config_window, text="Add Percent Complete", command=self.add_percent_complete_column)
            add_percent_complete_button.grid(row=8, column=2, padx=5, pady=5, sticky='ew')

            tk.Label(self.col_config_window, text="Phase ID Criteria:").grid(row=9, column=0, padx=5, pady=5, sticky='w')
            self.phase_id_criteria = tk.StringVar(self.col_config_window)
            if 'Phase ID' in self.df_expense.columns:
                phase_id_options = sorted([str(x) for x in self.df_expense['Phase ID'].unique()])
                if 'Supervisor' in phase_id_options:
                    self.phase_id_criteria.set('Supervisor')
                elif phase_id_options:
                    self.phase_id_criteria.set(phase_id_options[0])
                else:
                    self.phase_id_criteria.set("")
                self.phase_id_criteria_menu = tk.OptionMenu(self.col_config_window, self.phase_id_criteria, *phase_id_options)
                self.phase_id_criteria_menu.grid(row=9, column=1, padx=5, pady=5, sticky='ew')

            tk.Button(self.col_config_window, text="Save Configuration", command=self.save_app_configuration).grid(row=10, column=0, padx=5, pady=5, sticky='ew')
            tk.Button(self.col_config_window, text="Load Saved Configuration", command=self.load_app_configuration).grid(row=10, column=1, padx=5, pady=5, sticky='ew')

            tk.Button(self.col_config_window, text="Combine and Save", command=self.combine_and_save).grid(row=11, column=0, columnspan=2, pady=10)

            self.load_app_configuration()

        except Exception as e:
            messagebox.showerror("Error", f"Could not read files: {e}")

    def update_listbox(self):
        self.listbox.delete(0, tk.END)
        for col in self.original_columns:
            self.listbox.insert(tk.END, self.rename_map[col])

    def move_up(self):
        try:
            idx = self.listbox.curselection()[0]
            if idx > 0:
                self.original_columns.insert(idx - 1, self.original_columns.pop(idx))
                self.update_listbox()
                self.listbox.selection_set(idx - 1)
        except IndexError:
            pass

    def move_down(self):
        try:
            idx = self.listbox.curselection()[0]
            if idx < len(self.original_columns) - 1:
                self.original_columns.insert(idx + 1, self.original_columns.pop(idx))
                self.update_listbox()
                self.listbox.selection_set(idx + 1)
        except IndexError:
            pass

    def rename_column(self):
        try:
            idx = self.listbox.curselection()[0]
            original_col_name = self.original_columns[idx]
            old_display_name = self.rename_map[original_col_name]
            
            new_name = simpledialog.askstring("Rename Column", f"Enter new name for '{old_display_name}':", parent=self.col_config_window)
            if new_name:
                self.rename_map[original_col_name] = new_name
                self.update_listbox()
                self.listbox.selection_set(idx)
        except IndexError:
            messagebox.showwarning("Warning", "Please select a column to rename.", parent=self.col_config_window)

    def remove_column(self):
        try:
            idx = self.listbox.curselection()[0]
            self.original_columns.pop(idx)
            self.update_listbox()
        except IndexError:
            messagebox.showwarning("Warning", "Please select a column to remove.", parent=self.col_config_window)

    def select_key_columns(self):
        common_columns = sorted(list(set(self.df_expense.columns) & set(self.df_revenue.columns)))
        
        select_window = tk.Toplevel(self.col_config_window)
        select_window.title("Select Key Columns")
        
        listbox = tk.Listbox(select_window, selectmode=tk.MULTIPLE, width=40, height=15)
        for col in common_columns:
            listbox.insert(tk.END, col)
            if col in self.key_columns:
                listbox.selection_set(tk.END)
        listbox.pack(padx=10, pady=10)

        def on_ok():
            selected_indices = listbox.curselection()
            self.key_columns = [listbox.get(i) for i in selected_indices]
            self.key_columns_label.config(text=f"Keys: {self.key_columns}")
            select_window.destroy()

        tk.Button(select_window, text="OK", command=on_ok).pack(pady=5)

    def reset_key_columns(self):
        self.key_columns = ['Job ID', 'Cost Code ID', 'Phase ID']
        self.key_columns_label.config(text=f"Keys: {self.key_columns}")

    def select_supervisor_columns(self):
        all_columns = sorted(list(set(self.df_expense.columns) | set(self.df_revenue.columns)))
        
        select_window = tk.Toplevel(self.col_config_window)
        select_window.title("Select Supervisor Columns")
        
        tk.Label(select_window, text="Actual Expenses Column:").pack(padx=10, pady=5)
        actual_listbox = tk.Listbox(select_window, selectmode=tk.SINGLE, width=40, height=15)
        for col in all_columns:
            actual_listbox.insert(tk.END, col)
            if self.supervisor_columns and self.supervisor_columns.get('actual') == col:
                actual_listbox.selection_set(tk.END)
        actual_listbox.pack(padx=10, pady=5)

        tk.Label(select_window, text="Estimated Expenses Column:").pack(padx=10, pady=5)
        estimated_listbox = tk.Listbox(select_window, selectmode=tk.SINGLE, width=40, height=15)
        for col in all_columns:
            estimated_listbox.insert(tk.END, col)
            if self.supervisor_columns and self.supervisor_columns.get('estimated') == col:
                estimated_listbox.selection_set(tk.END)
        estimated_listbox.pack(padx=10, pady=5)

        def on_ok():
            actual_indices = actual_listbox.curselection()
            estimated_indices = estimated_listbox.curselection()
            if actual_indices and estimated_indices:
                self.supervisor_columns = {
                    'actual': actual_listbox.get(actual_indices[0]),
                    'estimated': estimated_listbox.get(estimated_indices[0])
                }
            else:
                self.supervisor_columns = {}
            self.supervisor_columns_label.config(text=f"Supervisor Columns: {self.supervisor_columns if self.supervisor_columns else 'None'}")
            select_window.destroy()

        tk.Button(select_window, text="OK", command=on_ok).pack(pady=5)

    def select_required_column(self):
        all_columns = sorted(list(set(self.df_expense.columns) | set(self.df_revenue.columns)))
        
        select_window = tk.Toplevel(self.col_config_window)
        select_window.title("Select Required Column")
        
        listbox = tk.Listbox(select_window, selectmode=tk.SINGLE, width=40, height=15)
        for col in all_columns:
            listbox.insert(tk.END, col)
            if col == self.required_column:
                listbox.selection_set(tk.END)
        listbox.pack(padx=10, pady=10)

        def on_ok():
            selected_indices = listbox.curselection()
            if selected_indices:
                self.required_column = listbox.get(selected_indices[0])
            else:
                self.required_column = None
            self.required_column_label.config(text=f"Required Column: {self.required_column if self.required_column else 'None'}")
            select_window.destroy()

        tk.Button(select_window, text="OK", command=on_ok).pack(pady=5)

    def reset_required_column(self):
        self.required_column = None
        self.required_column_label.config(text=f"Required Column: None")

    def add_percent_complete_column(self):
        if 'Percent Complete' not in self.original_columns:
            self.original_columns.append('Percent Complete')
            self.rename_map['Percent Complete'] = 'Percent Complete'
            self.update_listbox()

    def save_app_configuration(self):
        config_data = {
            'original_columns': self.original_columns,
            'rename_map': self.rename_map,
            'key_columns': self.key_columns,
            'required_column': self.required_column,
            'phase_id_criteria': self.phase_id_criteria.get()
        }
        with open(self.config_file, 'w') as f:
            json.dump(config_data, f, indent=4)
        messagebox.showinfo("Success", "Configuration saved.")

    def load_app_configuration(self):
        try:
            with open(self.config_file, 'r') as f:
                config_data = json.load(f)
                self.original_columns = config_data.get('original_columns', [])
                self.rename_map = config_data.get('rename_map', {})
                self.key_columns = config_data.get('key_columns', ['Job ID', 'Cost Code ID', 'Phase ID'])
                self.required_column = config_data.get('required_column', None)
                self.phase_id_criteria.set(config_data.get('phase_id_criteria', ""))
            if hasattr(self, 'listbox'):
                self.update_listbox()
            if hasattr(self, 'key_columns_label'):
                self.key_columns_label.config(text=f"Keys: {self.key_columns}")
            if hasattr(self, 'required_column_label'):
                self.required_column_label.config(text=f"Required Column: {self.required_column if self.required_column else 'None'}")
        except FileNotFoundError:
            messagebox.showerror("Error", "No saved configuration file found.")

    def reset_to_default(self):
        self.original_columns = self.default_original_columns.copy()
        self.rename_map = self.default_rename_map.copy()
        self.key_columns = ['Job ID', 'Cost Code ID', 'Phase ID']
        self.required_column = None
        self.update_listbox()
        self.key_columns_label.config(text=f"Keys: {self.key_columns}")
        self.required_column_label.config(text=f"Required Column: {self.required_column if self.required_column else 'None'}")
        messagebox.showinfo("Success", "Configuration reset to default.")
    def combine_and_save(self):
        try:
            if not self.key_columns:
                messagebox.showerror("Error", "Please select at least one key column.", parent=self.col_config_window)
                return

            if self.required_column:
                self.df_expense.dropna(subset=[self.required_column], inplace=True)
                self.df_revenue.dropna(subset=[self.required_column], inplace=True)

            if not all(col in self.df_expense.columns and col in self.df_revenue.columns for col in self.key_columns):
                 messagebox.showerror("Error", f"Both files must contain the key columns: {self.key_columns}", parent=self.col_config_window)
                 return

            merged_df = pd.merge(self.df_expense, self.df_revenue, on=self.key_columns, how='outer')

            if 'Percent Complete' in self.original_columns and self.supervisor_columns and self.supervisor_columns.get('actual') and self.supervisor_columns.get('estimated'):
                actual_col = self.supervisor_columns['actual']
                estimated_col = self.supervisor_columns['estimated']
                if actual_col in merged_df.columns and estimated_col in merged_df.columns:
                    merged_df['Percent Complete'] = pd.to_numeric(merged_df[actual_col], errors='coerce') / pd.to_numeric(merged_df[estimated_col], errors='coerce')
                    merged_df['Percent Complete'].fillna(0, inplace=True)
                    merged_df['Percent Complete'].replace([float('inf'), float('-inf')], 0, inplace=True)

            final_df = merged_df[self.original_columns]
            
            final_df = final_df.rename(columns=self.rename_map)

            with pd.ExcelWriter(self.output_file_path.get(), engine='openpyxl') as writer:
                final_df.to_excel(writer, index=False, sheet_name='Sheet1')

                # Get the workbook and worksheet objects
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']

                # Get the supervisor percent complete
                supervisor_percent_complete = 0
                if self.phase_id_criteria.get() and 'Phase ID' in final_df.columns and 'Percent Complete' in final_df.columns:
                    supervisor_row = final_df[final_df['Phase ID'] == self.phase_id_criteria.get()]
                    if not supervisor_row.empty:
                        supervisor_percent_complete = supervisor_row['Percent Complete'].iloc[0]

                # Define the cell fills
                green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

                # Find the column index for 'Percent Complete'
                percent_complete_col_index = None
                if 'Percent Complete' in final_df.columns:
                    for col in worksheet.iter_cols(min_row=1, max_row=1):
                        for cell in col:
                            if cell.value == 'Percent Complete':
                                percent_complete_col_index = cell.column
                                break
                        if percent_complete_col_index:
                            break
                
                    if percent_complete_col_index:
                        # Add conditional formatting rules
                        worksheet.conditional_formatting.add(f'{chr(ord('A') + percent_complete_col_index - 1)}2:{chr(ord('A') + percent_complete_col_index - 1)}{len(final_df) + 1}',
                                                           CellIsRule(operator='lessThan', formula=[supervisor_percent_complete], fill=green_fill))
                        worksheet.conditional_formatting.add(f'{chr(ord('A') + percent_complete_col_index - 1)}2:{chr(ord('A') + percent_complete_col_index - 1)}{len(final_df) + 1}',
                                                           CellIsRule(operator='greaterThan', formula=[supervisor_percent_complete], fill=red_fill))

                # Save a CSV for debugging
                csv_output_path = self.output_file_path.get().replace('.xlsx', '.csv')
                final_df.to_csv(csv_output_path, index=False)

            messagebox.showinfo("Success", f"Combined file saved to {self.output_file_path.get()}")
            self.col_config_window.destroy()

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred during merging: {e}", parent=self.col_config_window)

if __name__ == "__main__":
    root = tk.Tk()
    app = ProExcelMergerApp(root)
    root.mainloop()